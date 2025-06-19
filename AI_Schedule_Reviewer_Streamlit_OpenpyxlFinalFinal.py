
# 📌 Streamlit AI Schedule Reviewer - FINAL FINAL version with BytesIO buffer
# Handles Streamlit in-memory upload + openpyxl + embedded shapes robustly

import streamlit as st
import pandas as pd
import networkx as nx
import openai
from openpyxl import load_workbook
from io import BytesIO

st.title("🗂️ AI Schedule Reviewer (Final FINAL Openpyxl Version)")

uploaded_file = st.file_uploader("Upload your XER Toolkit export (.xlsx)", type=["xlsx"])

if uploaded_file:
    # Safe: wrap in BytesIO buffer
    bytes_data = uploaded_file.read()
    buffer = BytesIO(bytes_data)
    buffer.seek(0)

    wb = load_workbook(filename=buffer, data_only=True)

    st.write(f"Workbook Sheets: {wb.sheetnames}")

    sheets = {}

    if len(wb.sheetnames) > 3:
        for sheet_name in wb.sheetnames[3:]:
            try:
                ws = wb[sheet_name]
                data = []

                # Skip top 3 rows
                for row in ws.iter_rows(min_row=4, values_only=True):
                    data.append(row)

                # Drop fully empty rows
                data = [r for r in data if any(r)]

                if data:
                    df = pd.DataFrame(data[1:], columns=data[0])
                    sheets[sheet_name] = df
                else:
                    st.info(f"Skipped sheet '{sheet_name}' because it was empty after skip.")

            except Exception as e:
                st.warning(f"Skipped sheet '{sheet_name}' due to error: {e}")

    if not sheets:
        st.error("No valid data sheets found after skipping first 3 sheets. Please check your export format.")
    else:
        for name, df in sheets.items():
            st.subheader(f"Sheet: {name}")
            st.dataframe(df)

        activities = sheets.get("Activities")
        relationships = sheets.get("Relationships")

        if activities is not None and relationships is not None:
            st.subheader("Rule Checks")

            dangling = activities[activities['Predecessors'].isna() & activities['Successors'].isna()]
            st.write("Dangling Activities:", dangling)

            long_tasks = activities[activities['Duration'] > 30]
            st.write("Long Duration Activities (>30 days):", long_tasks)

            G = nx.DiGraph()
            for _, row in relationships.iterrows():
                G.add_edge(row['Predecessor'], row['Successor'])
            cycles = list(nx.simple_cycles(G))
            st.write("Logic Loops:", cycles)

            api_key = st.text_input("Enter your OpenAI API Key", type="password")

            if api_key and st.button("Generate AI Review"):
                openai.api_key = api_key
                issues_summary = f"Dangling: {len(dangling)}, Long tasks: {len(long_tasks)}, Loops: {len(cycles)}"
                response = openai.ChatCompletion.create(
                    model="gpt-4o",
                    messages=[
                        {"role": "system", "content": "You are a senior planner reviewing a construction schedule."},
                        {"role": "user", "content": f"Please review this schedule analysis: {issues_summary}. Provide recommendations in plain English."}
                    ]
                )
                st.subheader("AI Review Result")
                st.write(response['choices'][0]['message']['content'])
        else:
            st.warning("Activities or Relationships sheet not found among valid sheets!")
